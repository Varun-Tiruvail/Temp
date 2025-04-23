
# Add this import
from PySide6.QtGui import QIcon
import sys
import os
import pandas as pd
import polars as pl
from pathlib import Path
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QPushButton, 
                              QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, 
                              QStackedWidget, QFileDialog, QTableWidget, 
                              QDateEdit, QMessageBox, QTableWidgetItem, QCalendarWidget,
                              QFormLayout, QDialog, QTabWidget, QGroupBox)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QDropEvent, QDragEnterEvent, QPainterPath, QRegion, QTransform, QColor
from PySide6.QtWidgets import QPlainTextEdit, QHeaderView, QGraphicsDropShadowEffect
from PySide6.QtGui import QDoubleValidator
from io import StringIO  # Import StringIO for text conversion
# Add these imports at the top
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# Add these helper functions here, before any class definitions
def get_cell_value(value):
    """Convert cell value while preserving type and handling blanks"""
    if pd.isna(value):
        return ""  # Return empty string for blank cells
    if isinstance(value, (int, float)):
        if value.is_integer():
            return int(value)
        return value
    return str(value)

def populate_table_with_types(table_widget, df):
    """Populate QTableWidget while preserving data types"""
    if df is None or df.empty:
        table_widget.setRowCount(0)
        table_widget.setColumnCount(0)
        return
    
    table_widget.setRowCount(len(df))
    table_widget.setColumnCount(len(df.columns))
    table_widget.setHorizontalHeaderLabels(df.columns)
    
    for row in range(len(df)):
        for col in range(len(df.columns)):
            value = df.iloc[row, col]
            item = QTableWidgetItem()
            
            # Set alignment based on data type
            if isinstance(value, (int, float)):
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if pd.isna(value):  # Handle NaN/NA for numeric columns
                    item.setText("")
                else:
                    item.setData(Qt.DisplayRole, value)  # This preserves numeric type
            else:
                item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                item.setText("" if pd.isna(value) else str(value))
            
            table_widget.setItem(row, col, item)
    
    table_widget.resizeColumnsToContents()

def extract_table_data(table_widget):
    """Extract data from QTableWidget while preserving types"""
    data = []
    headers = []
    
    # Get headers
    for col in range(table_widget.columnCount()):
        headers.append(table_widget.horizontalHeaderItem(col).text())
    
    # Get data with types
    for row in range(table_widget.rowCount()):
        row_data = []
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            if item is None or not item.text():
                row_data.append("")  # Preserve blank cells
            else:
                # Try to convert to number if possible
                text = item.text()
                try:
                    # Check if it's an integer
                    if text.isdigit():
                        row_data.append(int(text))
                    # Check if it's a float
                    elif text.replace(".", "", 1).isdigit():
                        row_data.append(float(text))
                    else:
                        row_data.append(text)
                except ValueError:
                    row_data.append(text)
        data.append(row_data)
    
    return pd.DataFrame(data, columns=headers)


class TitleBar(QWidget):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.setObjectName("titleBar")
        self.setFixedHeight(45)  # Increased height for better visibility

        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 3, 5, 3)
        layout.setSpacing(3)
        
        # Title Label
        self.title_label = QLabel("RAPID - Rates Automation Projects Integrated Development")
        self.title_label.setObjectName("titleLabel")
        self.title_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        layout.addWidget(self.title_label)
        
        # Spacer
        layout.addStretch()
        
        # Window Controls
        self.minimize_btn = QPushButton("—")
        self.minimize_btn.setObjectName("minimizeButton")
        self.minimize_btn.setFixedSize(30, 25)
        self.minimize_btn.clicked.connect(parent.showMinimized)
        
        self.close_btn = QPushButton("✕")
        self.close_btn.setObjectName("closeButton")
        self.close_btn.setFixedSize(30, 25)
        self.close_btn.clicked.connect(parent.close)
        
        layout.addWidget(self.minimize_btn)
        layout.addWidget(self.close_btn)
        
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.parent.offset = event.globalPosition().toPoint() - self.parent.pos()
            
    def mouseMoveEvent(self, event):
        if self.parent.offset is not None and event.buttons() == Qt.LeftButton:
            self.parent.move(event.globalPosition().toPoint() - self.parent.offset)


class FramelessWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        
        # Create main container
        self.main_container = QWidget()
        self.main_container.setObjectName("mainContainer")
        self.setCentralWidget(self.main_container)
        
        # Create and apply shadow effect
        self.shadow = QGraphicsDropShadowEffect()
        self.shadow.setBlurRadius(20)
        self.shadow.setXOffset(0)
        self.shadow.setYOffset(0)
        self.shadow.setColor(QColor(0, 0, 0, 160))
        self.main_container.setGraphicsEffect(self.shadow)
        
        # Main vertical layout
        self.main_layout = QVBoxLayout(self.main_container)
        self.main_layout.setContentsMargins(10, 10, 10, 10)  # Add margins for shadow
        self.main_layout.setSpacing(0)
        
        # Add title bar
        self.title_bar = TitleBar(self)
        self.main_layout.addWidget(self.title_bar)
        
        # Content area
        self.content_area = QWidget()
        self.content_area.setObjectName("contentArea")
        self.main_layout.addWidget(self.content_area)

    def resizeEvent(self, event):
        # Update rounded corners when resizing
        path = QPainterPath()
        path.addRoundedRect(self.rect(), 10, 10)
        region = QRegion(path.toFillPolygon(QTransform()).toPolygon())
        self.setMask(region)
        super().resizeEvent(event)

class LoginWindow(QWidget):
    def __init__(self, excel_path="login_details.xlsx"):
        super().__init__()
        self.excel_path = excel_path
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("Login")
        self.setGeometry(300, 300, 300, 200)
        
        layout = QVBoxLayout()
        
        self.username_label = QLabel("Username:")
        self.username_input = QLineEdit()
        
        self.password_label = QLabel("Password:")
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        
        self.login_button = QPushButton("Login")
        self.login_button.clicked.connect(self.login)
        
        layout.addWidget(self.username_label)
        layout.addWidget(self.username_input)
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(self.login_button)
        
        self.setLayout(layout)
        
    def login(self):
        username = self.username_input.text()
        password = self.password_input.text()
        
        try:
            # Read Excel file with login credentials
            credentials_df = pd.read_excel(self.excel_path)
            
            # Check if credentials match
            valid_login = False
            for _, row in credentials_df.iterrows():
                if row['username'] == username and row['password'] == password:
                    valid_login = True
                    break
            
            if valid_login:
                self.hide()
                self.main_window = MainWindow()
                self.main_window.show()
            else:
                QMessageBox.warning(self, "Login Failed", "Invalid username or password")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Login error: {str(e)}")


# Main window after login
class MainWindow(FramelessWindow):
    # def __init__(self):
    #     super().__init__()
    #     self.init_ui()
    
    def __init__(self):
        super().__init__()
        self.init_ui()
        # self.setWindowTitle("RAPID - Rates Automation Projects Integrated Development")

    def init_ui(self):
        self.setGeometry(100, 100, 800, 600)
        content_layout = QVBoxLayout(self.content_area)
        self.ipv_button = QPushButton("IPV")
        self.reserves_button = QPushButton("Reserves")
        self.pruvals_button = QPushButton("Pruvals")
        self.other_tools_button = QPushButton("Other Tools")  # Use instance variable
        self.chatbot_button = QPushButton("AI Chatbot")

        # Create stacked widget before setting up connections
        self.stacked_widget = QStackedWidget()
        self.ipv_widget = IPVWidget()
        self.reserves_widget = ReservesWidget()
        self.other_tools_widget = OtherToolsWidget()  # Create instance here
        
        # Add widgets to stacked widget
        self.stacked_widget.addWidget(self.ipv_widget)
        self.stacked_widget.addWidget(self.reserves_widget)
        self.stacked_widget.addWidget(self.other_tools_widget)

        # Configure button layout
        button_layout = QHBoxLayout()
        buttons = [
            (self.ipv_button, self.ipv_widget),
            (self.reserves_button, self.reserves_widget),
            (self.pruvals_button, None),
            (self.other_tools_button, self.other_tools_widget),
            (self.chatbot_button, None)
        ]
        
        for btn, widget in buttons:
            btn.setCheckable(True)
            btn.setStyleSheet("""
                QPushButton {
                    padding: 8px 20px;
                    margin: 5px;
                }
                QPushButton:checked {
                    background-color: #4d5052;
                }
            """)
            button_layout.addWidget(btn)
            
            # Connect valid buttons to their widgets
            if widget:
                btn.clicked.connect(lambda _, w=widget: self.stacked_widget.setCurrentWidget(w))

        # Connect special buttons
        self.pruvals_button.clicked.connect(lambda: self.show_not_implemented("Pruvals"))
        self.chatbot_button.clicked.connect(lambda: self.show_not_implemented("AI Chatbot"))

        # Add layouts to main layout
        content_layout.addLayout(button_layout)
        content_layout .addWidget(self.stacked_widget)    
        
    def show_not_implemented(self, feature_name):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(f"{feature_name} feature is not yet ready to use.")
        msg.setInformativeText("Please try again later.")
        msg.setWindowTitle("Feature Coming Soon")
        msg.exec()


# IPV Widget
class IPVWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("IPV Categories")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        
        # Buttons
        self.xipv_button = QPushButton("XIPV")
        self.yipv_button = QPushButton("YIPV")
        
        # Connect buttons
        self.xipv_button.clicked.connect(self.open_xipv_window)
        self.yipv_button.clicked.connect(self.open_yipv_window)
        
        # Add widgets to layout
        layout.addWidget(title)
        layout2 = QHBoxLayout()
        layout2.addWidget(self.xipv_button)
        layout2.addWidget(self.yipv_button)
        layout.addLayout(layout2)

        
        self.setLayout(layout)
        
    def open_xipv_window(self):
        self.xipv_window = XIPVWindow()
        self.xipv_window.show()
        
    def open_yipv_window(self):
        self.yipv_window = YIPVWindow()
        self.yipv_window.show()


# Reserves Widget
class ReservesWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("Reserves Categories")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        
        # Buttons
        self.xreserves_button = QPushButton("XReserves")
        self.yreserves_button = QPushButton("YReserves")
        
        # Connect buttons
        self.xreserves_button.clicked.connect(self.open_xreserves_window)
        self.yreserves_button.clicked.connect(self.open_yreserves_window)
        
        # Add widgets to layout
        layout.addWidget(title)
        layout.addWidget(self.xreserves_button)
        layout.addWidget(self.yreserves_button)
        
        self.setLayout(layout)
        
    def open_xreserves_window(self):
        self.xreserves_window = XReservesWindow()
        self.xreserves_window.show()
        
    def open_yreserves_window(self):
        self.yreserves_window = YReservesWindow()
        self.yreserves_window.show()


# Updated XIPVWindow and related classes
class XIPVWindow(FramelessWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.data_entries = {}
        self.default_data_path = "FXDerivatives.xlsx"  # Path to default data Excel file
        
    def init_ui(self):
        self.setWindowTitle("XIPV")
        self.setGeometry(200, 200, 800, 600)
        
        # Central widget
        # central_widget = QWidget()
        # self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(self.content_area)
        
        # Title
        title = QLabel("XIPV Processing")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: bold;")
        
        # Process button
        self.process_button = QPushButton("Process Data")
        self.process_button.clicked.connect(self.start_process_sequence)
        
        # Table for displaying input data
        self.input_table = QTableWidget()
        self.input_table.setColumnCount(7)
        self.input_table.setHorizontalHeaderLabels([
            "File Path", "Date", "Adjustment 1", "Adjustment 2", 
            "Tables Imported", "Status", "Result"
        ])
        self.input_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        
        # Add widgets to layout
        main_layout.addWidget(title)
        main_layout.addWidget(self.process_button)
        main_layout.addWidget(self.input_table)
    
    def start_process_sequence(self):
        # Reset data for new processing
        self.data_entries = {
            'file_info': {},
            'tables': [None, None, None, None]  # Initialize with 4 None tables
        }
        
        # Show first dialog for file path, date, and adjustments
        self.show_file_info_dialog()
    
    def show_file_info_dialog(self):
        dialog = XIPVFileInfoDialog(self)
        if dialog.exec():
            # Store file info
            self.data_entries['file_info'] = {
                'file_path': dialog.file_path_input.text(),
                'date': dialog.date_input.date().toString("yyyy-MM-dd"),
                'adjustment1': dialog.adjustment1_input.text(),
                'adjustment2': dialog.adjustment2_input.text()
            }
            
            # Add data to table
            row_position = self.input_table.rowCount()
            self.input_table.insertRow(row_position)
            self.input_table.setItem(row_position, 0, QTableWidgetItem(self.data_entries['file_info']['file_path']))
            self.input_table.setItem(row_position, 1, QTableWidgetItem(self.data_entries['file_info']['date']))
            self.input_table.setItem(row_position, 2, QTableWidgetItem(self.data_entries['file_info']['adjustment1']))
            self.input_table.setItem(row_position, 3, QTableWidgetItem(self.data_entries['file_info']['adjustment2']))
            self.input_table.setItem(row_position, 4, QTableWidgetItem("0/4"))
            self.input_table.setItem(row_position, 5, QTableWidgetItem("Processing..."))
            
            # Start table import sequence with table 1 (mandatory)
            self.current_row = row_position
            self.show_table1_dialog()
    
    def show_table1_dialog(self):
        # Table 1 is mandatory
        dialog = XIPVTableDialog(self, table_number=1, is_mandatory=True)
        if dialog.exec():
            # Get table data and convert to polars DataFrame
            table_data = dialog.table_data
            if table_data:
                try:
                    # Parse the clipboard data into pandas DataFrame
                    pandas_df = pd.read_clipboard(sep='\t')
                    # Convert to polars
                    pl_df = pl.from_pandas(pandas_df)
                    
                    # Store the polars DataFrame
                    self.data_entries['tables'][0] = pl_df
                    
                    # Update table display
                    self.input_table.setItem(
                        self.current_row, 
                        4, 
                        QTableWidgetItem("1/4")
                    )
                    
                    # Now show dialog for remaining tables with default options
                    self.show_remaining_tables_dialog()
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to parse table data: {str(e)}")
                    # Try again with table 1
                    self.show_table1_dialog()
            else:
                # No data provided, try again
                QMessageBox.warning(self, "Warning", "No table data provided. Please paste data from Excel.")
                self.show_table1_dialog()
        else:
            # Dialog cancelled, update status
            self.input_table.setItem(self.current_row, 5, QTableWidgetItem("Cancelled"))

    def show_remaining_tables_dialog(self):
        # Load default data for remaining tables
        try:
            self.load_default_tables()
            
            # Show dialog for remaining tables
            dialog = XIPVRemainingTablesDialog(self, self.data_entries['tables'])
            if dialog.exec():
                # Update tables with any changes
                self.data_entries['tables'] = dialog.tables
                
                # Save any changed default tables to Excel
                self.save_default_tables(dialog.modified_tables)
                
                # Update table display
                self.input_table.setItem(
                    self.current_row, 
                    4, 
                    QTableWidgetItem("4/4")
                )
                
                # Process all data
                self.process_data()
            else:
                # Dialog cancelled, update status
                self.input_table.setItem(self.current_row, 5, QTableWidgetItem("Cancelled"))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error with default tables: {str(e)}")
            self.input_table.setItem(self.current_row, 5, QTableWidgetItem("Failed - Default tables error"))
    
    def load_default_tables(self):
        """Load default tables data from Excel file"""
        if os.path.exists(self.default_data_path):
            # Load tables 2-4 from Excel sheets
            for i in range(1, 4):  # Tables 2-4 (index 1-3)
                try:
                    pandas_df = pd.read_excel(self.default_data_path, sheet_name=f"Table{i+1}")
                    self.data_entries['tables'][i] = pl.from_pandas(pandas_df)
                except Exception as e:
                    print(f"Error loading default table {i+1}: {str(e)}")
                    # Create empty table as fallback
                    self.data_entries['tables'][i] = pl.DataFrame()
        else:
            # Create empty default tables
            for i in range(1, 4):
                self.data_entries['tables'][i] = pl.DataFrame()
            
            # Create default Excel file with empty sheets
            with pd.ExcelWriter(self.default_data_path) as writer:
                for i in range(1, 5):  # Create 4 sheets
                    pd.DataFrame().to_excel(writer, sheet_name=f"Table{i}", index=False)
    
    # def save_default_tables(self, modified_indices):
    #     """Save modified default tables back to Excel"""
    #     if not os.path.exists(self.default_data_path):
    #         # Create Excel file if it doesn't exist
    #         with pd.ExcelWriter(self.default_data_path) as writer:
    #             for i in range(4):
    #                 pd.DataFrame().to_excel(writer, sheet_name=f"Table{i+1}", index=False)
        
    #     # Only save tables that were modified
    #     if modified_indices:
    #         with pd.ExcelWriter(self.default_data_path, mode='a', if_sheet_exists='replace') as writer:
    #             for i in modified_indices:
    #                 if i > 0:  # Only save tables 2-4 (index 1-3)
    #                     # Convert polars to pandas for Excel writing
    #                     pandas_df = self.data_entries['tables'][i].to_pandas()
    #                     pandas_df.to_excel(writer, sheet_name=f"Table{i+1}", index=False)
    def save_default_tables(self, modified_indices):
        """Save modified default tables back to Excel"""
        if not modified_indices:
            return
        try:
            # Read existing data first
            existing_data = {}
            if os.path.exists(self.default_data_path):
                with pd.ExcelFile(self.default_data_path) as excel:
                    for sheet_name in excel.sheet_names:
                        existing_data[sheet_name] = pd.read_excel(excel, sheet_name=sheet_name)

            # Update modified sheets
            for i in modified_indices:
                if i > 0:  # Only save tables 2-4 (index 1-3)
                    sheet_name = f"Table{i+1}"
                    existing_data[sheet_name] = self.data_entries['tables'][i].to_pandas()

            # Write all sheets back to file
            with pd.ExcelWriter(self.default_data_path, engine='openpyxl') as writer:
                for sheet_name, df in existing_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save default tables: {str(e)}")
    
    def process_data(self):
        try:
            # Load data from file
            file_path = self.data_entries['file_info']['file_path']
            date = self.data_entries['file_info']['date']
            adjustment1 = float(self.data_entries['file_info']['adjustment1'])
            adjustment2 = float(self.data_entries['file_info']['adjustment2'])
            
            # Get all tables
            tables = self.data_entries['tables']
            
            if os.path.exists(file_path) and tables[0] is not None:
                # Determine file type and read into DataFrame
                file_ext = Path(file_path).suffix.lower()
                
                if file_ext == '.csv':
                    file_df = pl.read_csv(file_path)
                elif file_ext in ['.xlsx', '.xls']:
                    # Convert pandas to polars
                    pandas_df = pd.read_excel(file_path)
                    file_df = pl.from_pandas(pandas_df)
                else:
                    QMessageBox.warning(self, "Error", "Unsupported file format")
                    self.input_table.setItem(self.current_row, 5, QTableWidgetItem("Failed - Bad file format"))
                    return
                
                # ====== ADD YOUR CUSTOM FUNCTION HERE ======
                # Process using all dataframes and parameters
                # Example: 
                # result = process_xipv_data(
                #     file_df, 
                #     date, 
                #     adjustment1, 
                #     adjustment2,
                #     tables[0],  # Table 1
                #     tables[1],  # Table 2
                #     tables[2],  # Table 3
                #     tables[3]   # Table 4
                # )
                result = 0
                # ============================================
                
                # Update table
                self.input_table.setItem(self.current_row, 5, QTableWidgetItem("Completed"))
                self.input_table.setItem(self.current_row, 6, QTableWidgetItem(str(result)))
                
                QMessageBox.information(self, "Success", f"Data processed successfully. Result: {result}")
            else:
                QMessageBox.warning(self, "Error", "File not found or first table missing")
                self.input_table.setItem(self.current_row, 5, QTableWidgetItem("Failed - Missing data"))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Processing error: {str(e)}")
            self.input_table.setItem(self.current_row, 5, QTableWidgetItem(f"Failed - {str(e)[:20]}..."))


# Modified Dialog for XIPV table input (supports mandatory/optional)
class XIPVTableDialog(QDialog):
    def __init__(self, parent=None, table_number=1, is_mandatory=False):
        super().__init__(parent)
        self.setWindowTitle(f"XIPV Table {table_number} Input")
        self.setMinimumSize(600, 400)
        self.table_number = table_number
        self.is_mandatory = is_mandatory
        self.table_data = None
        
        layout = QVBoxLayout()
        
        # Instructions
        instructions = QLabel(
            f"Paste Table {table_number} data from Excel (Ctrl+V).\n"
            "Make sure to copy the entire table including headers."
        )
        if is_mandatory:
            instructions.setText(instructions.text() + "\n(This table is mandatory)")
        instructions.setStyleSheet("font-weight: bold;")
        
        # Text area for pasted data
        self.data_text = QPlainTextEdit()
        self.data_text.setPlaceholderText("Paste Excel data here...")
        
        # Preview area
        preview_label = QLabel("Data Preview:")
        self.preview_table = QTableWidget()
        
        # Connect paste event
        self.data_text.textChanged.connect(self.update_preview)
        
        # Buttons
        button_box = QHBoxLayout()
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.ok_button = QPushButton("Next")
        self.ok_button.clicked.connect(self.accept)
        
        button_box.addWidget(self.cancel_button)
        button_box.addWidget(self.ok_button)
        
        # Add to layout
        layout.addWidget(instructions)
        layout.addWidget(self.data_text)
        layout.addWidget(preview_label)
        layout.addWidget(self.preview_table)
        layout.addLayout(button_box)
        
        self.setLayout(layout)

    def update_preview(self):
        text = self.data_text.toPlainText()
        if text:
            try:
                # Convert the pasted text into a file-like object
                text_io = StringIO(text)

                # Try to parse the text as a DataFrame
                df = pd.read_csv(text_io, sep="\t")

                # Update preview table
                self.preview_table.setRowCount(min(5, len(df)))
                self.preview_table.setColumnCount(len(df.columns))
                self.preview_table.setHorizontalHeaderLabels(df.columns)

                # Fill preview data (first 5 rows)
                for i in range(min(5, len(df))):
                    for j in range(len(df.columns)):
                        item = QTableWidgetItem(str(df.iloc[i, j]))
                        self.preview_table.setItem(i, j, item)

                # Resize columns to content
                self.preview_table.resizeColumnsToContents()

                # Store the data
                self.table_data = text
            except Exception as e:
                # Handle parsing failures
                self.preview_table.setRowCount(0)
                self.preview_table.setColumnCount(0)
                self.table_data = None
                print(f"Error parsing data: {e}")
        else:
            # Clear preview if no text
            self.preview_table.setRowCount(0)
            self.preview_table.setColumnCount(0)
            self.table_data = None


# New dialog for managing remaining tables
class XIPVRemainingTablesDialog(QDialog):

    def __init__(self, parent=None, tables=None):
        super().__init__(parent)
        self.setWindowTitle("Remaining Tables")
        self.setMinimumSize(800, 600)
        self.tables = tables if tables else [None, None, None, None]
        self.modified_tables = set()  # Track which tables have been modified
        
        # Main layout
        main_layout = QVBoxLayout()
        
        # Instructions
        instructions = QLabel(
            "Table 1 has been imported. Default data is loaded for Tables 2-4.\n"
            "You can view and modify the default data or import new data."
        )
        instructions.setStyleSheet("font-weight: bold;")
        main_layout.addWidget(instructions)
        
        # Create tab widget for tables
        self.tab_widget = QTabWidget()
        
        # Add tabs for tables 2-4
        for i in range(1, 4):
            tab = self.create_table_tab(i+1, self.tables[i])
            self.tab_widget.addTab(tab, f"Table {i+1}")
        
        main_layout.addWidget(self.tab_widget)
        
        # Buttons
        button_box = QHBoxLayout()
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.ok_button = QPushButton("Finish")
        self.ok_button.clicked.connect(self.accept)
        
        button_box.addWidget(self.cancel_button)
        button_box.addWidget(self.ok_button)
        main_layout.addLayout(button_box)
        
        self.setLayout(main_layout)
    
    def create_table_tab(self, table_num, table_data):
        """Create a tab for a table"""
        tab_widget = QWidget()
        tab_layout = QVBoxLayout()
        
        # Button to import new data
        import_button = QPushButton(f"Import New Data for Table {table_num}")
        import_button.clicked.connect(lambda: self.import_new_data(table_num-1))
        tab_layout.addWidget(import_button)
        
        # Table widget to display/edit data
        table_widget = QTableWidget()
        self.populate_table_widget(table_widget, table_data)
        table_widget.itemChanged.connect(lambda: self.handle_table_edit(table_num-1, table_widget))
        
        tab_layout.addWidget(QLabel("Default Data (editable):"))
        tab_layout.addWidget(table_widget)
        
        # Store reference to table widget
        setattr(self, f"table{table_num}_widget", table_widget)
        
        tab_widget.setLayout(tab_layout)
        return tab_widget
    
    # def populate_table_widget(self, table_widget, table_data):
    #     """Populate a QTableWidget with data from a polars DataFrame"""
    #     if table_data is None or table_data.is_empty():
    #         table_widget.setRowCount(0)
    #         table_widget.setColumnCount(0)
    #         return
        
    #     # Convert polars to pandas for easier handling
    #     pandas_df = table_data.to_pandas()
        
    #     # Set table dimensions
    #     table_widget.setRowCount(len(pandas_df))
    #     table_widget.setColumnCount(len(pandas_df.columns))
    #     table_widget.setHorizontalHeaderLabels(pandas_df.columns)
        
    #     # Fill data
    #     for i in range(len(pandas_df)):
    #         for j in range(len(pandas_df.columns)):
    #             value = pandas_df.iloc[i, j]
    #             item = QTableWidgetItem(str(value))
    #             table_widget.setItem(i, j, item)
        
    #     # Resize columns to content
    #     table_widget.resizeColumnsToContents()
    
    def populate_table_widget(self, table_widget, table_data):
        """Populate a QTableWidget with data from a polars DataFrame"""
        if table_data is None or table_data.is_empty():
            table_widget.setRowCount(0)
            table_widget.setColumnCount(0)
            return
        
        # Convert polars to pandas for easier handling
        pandas_df = table_data.to_pandas()
        populate_table_with_types(table_widget, pandas_df)

    def import_new_data(self, table_index):
        """Import new data for a table"""
        dialog = XIPVTableDialog(self, table_number=table_index+1)
        if dialog.exec():
            table_data = dialog.table_data
            if table_data:
                try:
                    # Parse the clipboard data into pandas DataFrame
                    pandas_df = pd.read_clipboard(sep='\t')
                    # Convert to polars
                    pl_df = pl.from_pandas(pandas_df)
                    
                    # Update the table
                    self.tables[table_index] = pl_df
                    
                    # Update UI
                    table_widget = getattr(self, f"table{table_index+1}_widget")
                    self.populate_table_widget(table_widget, pl_df)
                    
                    # Mark as modified
                    self.modified_tables.add(table_index)
                    
                    QMessageBox.information(self, "Success", f"New data imported for Table {table_index+1}")
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to parse table data: {str(e)}")
    
    # def handle_table_edit(self, table_index, table_widget):
    #     """Handle edits to the table widget"""
    #     try:
    #         # Get headers
    #         headers = []
    #         for j in range(table_widget.columnCount()):
    #             headers.append(table_widget.horizontalHeaderItem(j).text())
            
    #         # Get all data
    #         data = []
    #         for i in range(table_widget.rowCount()):
    #             row_data = []
    #             for j in range(table_widget.columnCount()):
    #                 item = table_widget.item(i, j)
    #                 value = item.text() if item else ""
    #                 row_data.append(value)
    #             data.append(row_data)
            
    #         # Create pandas DataFrame and convert to polars
    #         pandas_df = pd.DataFrame(data, columns=headers)
    #         pl_df = pl.from_pandas(pandas_df)
            
    #         # Update the table
    #         self.tables[table_index] = pl_df
            
    #         # Mark as modified
    #         self.modified_tables.add(table_index)
    #     except Exception as e:
    #         print(f"Error updating table data: {str(e)}")

    def handle_table_edit(self, table_index, table_widget):
        """Handle edits to the table widget preserving types"""
        try:
            # Extract data preserving types
            pandas_df = extract_table_data(table_widget)
            
            # Convert to polars
            pl_df = pl.from_pandas(pandas_df)
            
            # Update the table
            self.tables[table_index] = pl_df
            
            # Mark as modified
            self.modified_tables.add(table_index)
        except Exception as e:
            print(f"Error updating table data: {str(e)}")

# Dialog for XIPV file info input (first popup)
class XIPVFileInfoDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("XIPV File Information")
        self.setMinimumWidth(400)
        
        layout = QFormLayout()
        
        # File path input
        self.file_path_input = QLineEdit()
        self.browse_button = QPushButton("Browse")
        self.browse_button.clicked.connect(self.browse_file)
        
        file_layout = QHBoxLayout()
        file_layout.addWidget(self.file_path_input)
        file_layout.addWidget(self.browse_button)
        
        # Date input
        self.date_input = QDateEdit()
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setCalendarPopup(True)
        
        # Adjustment inputs
        self.adjustment1_input = QLineEdit()
        self.adjustment2_input = QLineEdit()
        
        # Set placeholders and validators
        self.adjustment1_input.setPlaceholderText("Enter numeric value")
        self.adjustment2_input.setPlaceholderText("Enter numeric value")
        validator = QDoubleValidator()
        self.adjustment1_input.setValidator(validator)
        self.adjustment2_input.setValidator(validator)
        
        # Add to form layout
        layout.addRow("File Path:", file_layout)
        layout.addRow("Date:", self.date_input)
        layout.addRow("Adjustment 1:", self.adjustment1_input)
        layout.addRow("Adjustment 2:", self.adjustment2_input)
        
        # Buttons
        button_box = QHBoxLayout()
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.ok_button = QPushButton("Next")
        self.ok_button.clicked.connect(self.validate_and_accept)
        
        button_box.addWidget(self.cancel_button)
        button_box.addWidget(self.ok_button)
        
        layout.addRow("", button_box)
        
        self.setLayout(layout)
    
    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select File", "", "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            self.file_path_input.setText(file_path)
    
    def validate_and_accept(self):
        # Check if all fields are filled
        if not self.file_path_input.text():
            QMessageBox.warning(self, "Warning", "Please select a file.")
            return
        
        if not self.adjustment1_input.text() or not self.adjustment2_input.text():
            QMessageBox.warning(self, "Warning", "Please enter both adjustment values.")
            return
        
        self.accept()


# YIPV Window
class YIPVWindow(FramelessWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("YIPV")
        self.setGeometry(200, 200, 600, 400)
        
        # Central widget
        # central_widget = QWidget()
        # self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(self.content_area)
        
        # Title
        title = QLabel("YIPV Processing")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: bold;")
        
        # Create a horizontal layout for the extraction and broil file buttons
        extraction_buttons_layout = QHBoxLayout()
        
        # Extract button
        self.extract_button = QPushButton("Extract from the Cube")
        self.extract_button.clicked.connect(self.extract_from_cube)
        
        # Generate broil file button
        self.broil_button = QPushButton("Generate Broil File")
        self.broil_button.clicked.connect(self.generate_broil_file)
        
        # Add buttons to horizontal layout
        extraction_buttons_layout.addWidget(self.extract_button)
        extraction_buttons_layout.addWidget(self.broil_button)
        
        # Create drag and drop area
        self.drag_drop_area = FileDragDropWidget()
        
        # Calculate button - now below drag and drop area
        self.calculate_button = QPushButton("Calculate")
        self.calculate_button.clicked.connect(self.calculate)
        
        # Add widgets and layouts to main layout
        main_layout.addWidget(title)
        main_layout.addLayout(extraction_buttons_layout)  # Add the horizontal layout
        main_layout.addWidget(QLabel("Drag and drop file here:"))
        main_layout.addWidget(self.drag_drop_area)
        main_layout.addWidget(self.calculate_button)  # Moved below drag and drop area
    
    def extract_from_cube(self):
        date, ok = self.get_date_input()
        if ok:
            # ====== ADD YOUR EXTRACT FUNCTION HERE ======
            # Replace this comment with your extract function
            # Example: result = extract_from_cube(date)
            # Currently just returning None as per requirements
            result = None
            # ============================================
            
            QMessageBox.information(self, "Extraction", f"Extracted data for {date}")
    
    def generate_broil_file(self):
        date, ok = self.get_date_input()
        if ok:
            # ====== ADD YOUR BROIL FILE GENERATION FUNCTION HERE ======
            # Replace this comment with your broil file generation function
            # Example: result = generate_broil_file(date)
            # Currently just showing a message
            # ============================================
            
            QMessageBox.information(self, "Broil File Generation", f"Generated broil file for {date}")
    
    def calculate(self):
        if hasattr(self.drag_drop_area, 'file_path') and self.drag_drop_area.file_path:
            file_path = self.drag_drop_area.file_path
            
            # ====== ADD YOUR CALCULATE FUNCTION HERE ======
            # Replace this comment with your calculate function
            # Example: result = calculate_yipv(file_path)
            # Currently just returning 0 as per requirements
            result = 0
            # =============================================
            
            QMessageBox.information(self, "Calculation", f"Calculation complete. Result: {result}")
        else:
            QMessageBox.warning(self, "Error", "Please drag and drop a file first")
    
    def get_date_input(self):
        date_dialog = QDialog(self)
        date_dialog.setWindowTitle("Enter Date")
        
        layout = QVBoxLayout()
        
        date_edit = QDateEdit()
        date_edit.setDate(QDate.currentDate())
        date_edit.setCalendarPopup(True)
        
        buttons = QHBoxLayout()
        cancel_button = QPushButton("Cancel")
        ok_button = QPushButton("OK")
        
        buttons.addWidget(cancel_button)
        buttons.addWidget(ok_button)
        
        layout.addWidget(QLabel("Select date:"))
        layout.addWidget(date_edit)
        layout.addLayout(buttons)
        
        date_dialog.setLayout(layout)
        
        ok_button.clicked.connect(date_dialog.accept)
        cancel_button.clicked.connect(date_dialog.reject)
        
        result = date_dialog.exec()
        return date_edit.date().toString("yyyy-MM-dd"), result == QDialog.Accepted


# File drag and drop widget
class FileDragDropWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.file_path = None
        
    def init_ui(self):
        self.setAcceptDrops(True)
        self.setMinimumHeight(100)
        self.setStyleSheet("border: 2px dashed #aaa; border-radius: 5px;")
        
        layout = QVBoxLayout()
        
        self.label = QLabel("Drop file here")
        self.label.setAlignment(Qt.AlignCenter)
        
        layout.addWidget(self.label)
        self.setLayout(layout)
    
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
    
    def dropEvent(self, event: QDropEvent):
        if event.mimeData().hasUrls():
            url = event.mimeData().urls()[0]
            self.file_path = url.toLocalFile()
            self.label.setText(f"File: {os.path.basename(self.file_path)}")
            event.acceptProposedAction()


class XReservesWindow(FramelessWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("XReserves")
        self.setGeometry(200, 200, 400, 200)
        
        # Central widget
        # central_widget = QWidget()
        # self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(self.content_area)
        
        # Title
        title = QLabel("XReserves Processing")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: bold;")
        
        # Buttons
        self.spreads_button = QPushButton("Calculate Spreads")
        self.spreads_button.clicked.connect(self.calculate_spreads)
        
        self.reserve_button = QPushButton("Allocate Reserve")
        self.reserve_button.clicked.connect(self.show_reserve_window)
        
        # Add widgets to layout
        main_layout.addWidget(title)
        main_layout.addWidget(self.spreads_button)
        main_layout.addWidget(self.reserve_button)
        
    def calculate_spreads(self):
    # Show first file dialog
        first_dialog = FirstFileDialog(self)
        if first_dialog.exec():
            first_file = first_dialog.file_path
            date = first_dialog.date_input.date().toString("yyyy-MM-dd")
            
            # Show second file dialog
            second_dialog = SecondFileDialog(self)
            if second_dialog.exec():
                second_file = second_dialog.file_path
                
                # Your spreads calculation function will go here
                QMessageBox.information(self, "Processing", 
                                    f"Calculating spreads...\n"
                                    f"First file: {os.path.basename(first_file)}\n"
                                    f"Second file: {os.path.basename(second_file)}\n"
                                    f"Date: {date}")
    def show_reserve_window(self):
        self.reserve_window = XReservesAllocationWindow()
        self.reserve_window.show()

# Class for XReserves file info dialog (similar to XIPV)
class SpreadsFileDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Load Files for Spreads Calculation")
        self.setMinimumWidth(600)
        self.setMinimumHeight(400)
        self.first_file_path = None
        self.second_file_path = None
        
        layout = QVBoxLayout()
        
        # First file section
        first_file_group = QGroupBox("First File")
        first_file_layout = QVBoxLayout()
        
        # First drag-drop area
        self.first_drag_drop = FileDragDropWidget()
        
        # First browse button
        self.first_browse_button = QPushButton("Browse")
        self.first_browse_button.clicked.connect(lambda: self.browse_file("first"))
        
        first_file_layout.addWidget(self.first_drag_drop)
        first_file_layout.addWidget(self.first_browse_button)
        first_file_group.setLayout(first_file_layout)
        
        # Second file section
        second_file_group = QGroupBox("Second File")
        second_file_layout = QVBoxLayout()
        
        # Second drag-drop area
        self.second_drag_drop = FileDragDropWidget()
        
        # Second browse button
        self.second_browse_button = QPushButton("Browse")
        self.second_browse_button.clicked.connect(lambda: self.browse_file("second"))
        
        second_file_layout.addWidget(self.second_drag_drop)
        second_file_layout.addWidget(self.second_browse_button)
        second_file_group.setLayout(second_file_layout)
        
        # Date input
        date_layout = QHBoxLayout()
        self.date_input = QDateEdit()
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setCalendarPopup(True)
        date_layout.addWidget(QLabel("Date:"))
        date_layout.addWidget(self.date_input)
        
        # Buttons
        button_box = QHBoxLayout()
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.calculate_button = QPushButton("Calculate")
        self.calculate_button.clicked.connect(self.validate_and_accept)
        
        button_box.addWidget(self.cancel_button)
        button_box.addWidget(self.calculate_button)
        
        # Add all widgets to main layout
        layout.addWidget(first_file_group)
        layout.addWidget(second_file_group)
        layout.addLayout(date_layout)
        layout.addLayout(button_box)
        
        self.setLayout(layout)
    
    def browse_file(self, file_type):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select File", "", "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            if file_type == "first":
                self.first_file_path = file_path
                self.first_drag_drop.label.setText(f"File: {os.path.basename(file_path)}")
            else:
                self.second_file_path = file_path
                self.second_drag_drop.label.setText(f"File: {os.path.basename(file_path)}")
    
    def validate_and_accept(self):
        self.first_file_path = getattr(self.first_drag_drop, 'file_path', None)
        self.second_file_path = getattr(self.second_drag_drop, 'file_path', None)
        
        if not self.first_file_path or not self.second_file_path:
            QMessageBox.warning(self, "Warning", "Please select both files.")
            return
        
        self.accept()


class XReservesAllocationWindow(FramelessWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.data_entries = {
            'tables': [None] * 7  # 2 mandatory + 5 optional tables
        }
        self.default_data_path = "FXMMResrveAllocation.xlsx"
        
    def init_ui(self):
        self.setWindowTitle("XReserves Allocation")
        self.setGeometry(200, 200, 800, 600)
        
        # Central widget
        # central_widget = QWidget()
        # self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(self.content_area)
        
        # Title
        title = QLabel("XReserves Allocation")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: bold;")
        
        # Process button
        self.process_button = QPushButton("Start Process")
        self.process_button.clicked.connect(self.start_process_sequence)
        
        # Table for displaying status
        self.input_table = QTableWidget()
        self.input_table.setColumnCount(3)
        self.input_table.setHorizontalHeaderLabels([
            "Tables", "Status", "Result"
        ])
        self.input_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        
        # Add widgets to layout
        main_layout.addWidget(title)
        main_layout.addWidget(self.process_button)
        main_layout.addWidget(self.input_table)
    
    def start_process_sequence(self):
        # Reset data
        self.data_entries['tables'] = [None] * 7
        
        # Show first mandatory table dialog
        self.current_row = self.input_table.rowCount()
        self.input_table.insertRow(self.current_row)
        self.input_table.setItem(self.current_row, 0, QTableWidgetItem("Mandatory Tables"))
        self.input_table.setItem(self.current_row, 1, QTableWidgetItem("Processing..."))
        
        self.show_mandatory_table1_dialog()
    
    def show_mandatory_table1_dialog(self):
        dialog = XReservesTableDialog(self, table_number=1, is_mandatory=True)
        if dialog.exec():
            if dialog.table_data:
                try:
                    pandas_df = pd.read_clipboard(sep='\t')
                    pl_df = pl.from_pandas(pandas_df)
                    self.data_entries['tables'][0] = pl_df
                    self.show_mandatory_table2_dialog()
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to parse table data: {str(e)}")
                    self.show_mandatory_table1_dialog()
            else:
                QMessageBox.warning(self, "Warning", "No table data provided")
                self.show_mandatory_table1_dialog()
        else:
            self.input_table.setItem(self.current_row, 1, QTableWidgetItem("Cancelled"))
    
    def show_mandatory_table2_dialog(self):
        dialog = XReservesTableDialog(self, table_number=2, is_mandatory=True)
        if dialog.exec():
            if dialog.table_data:
                try:
                    pandas_df = pd.read_clipboard(sep='\t')
                    pl_df = pl.from_pandas(pandas_df)
                    self.data_entries['tables'][1] = pl_df
                    self.show_remaining_tables_dialog()
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to parse table data: {str(e)}")
                    self.show_mandatory_table2_dialog()
            else:
                QMessageBox.warning(self, "Warning", "No table data provided")
                self.show_mandatory_table2_dialog()
        else:
            self.input_table.setItem(self.current_row, 1, QTableWidgetItem("Cancelled"))
    
    def show_remaining_tables_dialog(self):
        try:
            # Load default data for remaining tables
            self.load_default_tables()
            
            # Show dialog for remaining tables
            dialog = XReservesRemainingTablesDialog(self, self.data_entries['tables'][2:])
            if dialog.exec():
                # Update tables with any changes
                for i, table in enumerate(dialog.tables):
                    self.data_entries['tables'][i+2] = table
                
                # Save modified default tables
                self.save_default_tables(dialog.modified_tables)
                
                # Update status
                self.input_table.setItem(self.current_row, 1, QTableWidgetItem("Completed"))
                
                # Process the allocation
                self.process_allocation()
            else:
                self.input_table.setItem(self.current_row, 1, QTableWidgetItem("Cancelled"))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error with tables: {str(e)}")
            self.input_table.setItem(self.current_row, 1, QTableWidgetItem("Failed"))
    
    # def load_default_tables(self):
    #     """Load default data for optional tables from Excel file"""
    #     if os.path.exists(self.default_data_path):
    #         for i in range(5):  # 5 optional tables
    #             try:
    #                 pandas_df = pd.read_excel(self.default_data_path, sheet_name=f"Table{i+3}")
    #                 self.data_entries['tables'][i+2] = pl.from_pandas(pandas_df)
    #             except Exception as e:
    #                 print(f"Error loading default table {i+3}: {str(e)}")
    #                 self.data_entries['tables'][i+2] = pl.DataFrame()
    #     else:
    #         for i in range(5):
    #             self.data_entries['tables'][i+2] = pl.DataFrame()
    
    # def save_default_tables(self, modified_indices):
    #     """Save modified default tables back to Excel"""
    #     if modified_indices and os.path.exists(self.default_data_path):
    #         with pd.ExcelWriter(self.default_data_path, mode='a', if_sheet_exists='replace') as writer:
    #             for i in modified_indices:
    #                 pandas_df = self.data_entries['tables'][i+2].to_pandas()
    #                 pandas_df.to_excel(writer, sheet_name=f"Table{i+3}", index=False)
    def load_default_tables(self):
        """Load default data for optional tables from Excel file"""
        if os.path.exists(self.default_data_path):
            try:
                with pd.ExcelFile(self.default_data_path) as excel:
                    for i in range(5):  # 5 optional tables
                        sheet_name = f"Table{i+3}"
                        if sheet_name in excel.sheet_names:
                            pandas_df = pd.read_excel(excel, sheet_name=sheet_name)
                            self.data_entries['tables'][i+2] = pl.from_pandas(pandas_df)
                        else:
                            # Create empty table if sheet doesn't exist
                            self.data_entries['tables'][i+2] = pl.DataFrame()
            except Exception as e:
                print(f"Error loading default tables: {str(e)}")
                # Create empty tables as fallback
                for i in range(5):
                    self.data_entries['tables'][i+2] = pl.DataFrame()
        else:
            # Create empty default tables and Excel file
            for i in range(5):
                self.data_entries['tables'][i+2] = pl.DataFrame()
            
            # Create new Excel file with empty sheets
            with pd.ExcelWriter(self.default_data_path, engine='openpyxl') as writer:
                for i in range(5):
                    pd.DataFrame().to_excel(writer, sheet_name=f"Table{i+3}", index=False)

    def save_default_tables(self, modified_indices):
        """Save modified default tables back to Excel"""
        if not modified_indices:
            return

        try:
            # Read existing data first
            existing_data = {}
            if os.path.exists(self.default_data_path):
                with pd.ExcelFile(self.default_data_path) as excel:
                    for sheet_name in excel.sheet_names:
                        existing_data[sheet_name] = pd.read_excel(excel, sheet_name=sheet_name)

            # Update modified sheets
            for i in modified_indices:
                sheet_name = f"Table{i+3}"
                existing_data[sheet_name] = self.data_entries['tables'][i+2].to_pandas()

            # Write all sheets back to file
            with pd.ExcelWriter(self.default_data_path, engine='openpyxl') as writer:
                for sheet_name, df in existing_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save default tables: {str(e)}")
    
    def process_allocation(self):
        try:
            # Your allocation processing function will go here
            # You can access all tables through self.data_entries['tables']
            QMessageBox.information(self, "Processing", "Processing allocation with all tables")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Processing error: {str(e)}")

# Class for XReserves table dialog (similar to XIPV)
class XReservesTableDialog(QDialog):
    def __init__(self, parent=None, table_number=1, is_mandatory=False):
        super().__init__(parent)
        self.setWindowTitle(f"XReserves Table {table_number} Input")
        self.setMinimumSize(600, 400)
        self.table_number = table_number
        self.is_mandatory = is_mandatory
        self.table_data = None
        
        layout = QVBoxLayout()
        
        # Instructions
        instructions = QLabel(
            f"Paste Table {table_number} data from Excel (Ctrl+V).\n"
            "Make sure to copy the entire table including headers."
        )
        if is_mandatory:
            instructions.setText(instructions.text() + "\n(This table is mandatory)")
        instructions.setStyleSheet("font-weight: bold;")
        
        # Text area for pasted data
        self.data_text = QPlainTextEdit()
        self.data_text.setPlaceholderText("Paste Excel data here...")
        
        # Preview area
        preview_label = QLabel("Data Preview:")
        self.preview_table = QTableWidget()
        
        # Connect paste event
        self.data_text.textChanged.connect(self.update_preview)
        
        # Buttons
        button_box = QHBoxLayout()
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.ok_button = QPushButton("Next")
        self.ok_button.clicked.connect(self.accept)
        
        button_box.addWidget(self.cancel_button)
        button_box.addWidget(self.ok_button)
        
        # Add to layout
        layout.addWidget(instructions)
        layout.addWidget(self.data_text)
        layout.addWidget(preview_label)
        layout.addWidget(self.preview_table)
        layout.addLayout(button_box)
        
        self.setLayout(layout)

    def update_preview(self):
        text = self.data_text.toPlainText()
        if text:
            try:
                # Convert the pasted text into a file-like object
                text_io = StringIO(text)

                # Try to parse the text as a DataFrame
                df = pd.read_csv(text_io, sep="\t")

                # Update preview table
                self.preview_table.setRowCount(min(5, len(df)))
                self.preview_table.setColumnCount(len(df.columns))
                self.preview_table.setHorizontalHeaderLabels(df.columns)

                # Fill preview data (first 5 rows)
                for i in range(min(5, len(df))):
                    for j in range(len(df.columns)):
                        item = QTableWidgetItem(str(df.iloc[i, j]))
                        self.preview_table.setItem(i, j, item)

                # Resize columns to content
                self.preview_table.resizeColumnsToContents()

                # Store the data
                self.table_data = text
            except Exception as e:
                # Handle parsing failures
                self.preview_table.setRowCount(0)
                self.preview_table.setColumnCount(0)
                self.table_data = None
                print(f"Error parsing data: {e}")
        else:
            # Clear preview if no text
            self.preview_table.setRowCount(0)
            self.preview_table.setColumnCount(0)
            self.table_data = None

# Class for remaining XReserves tables dialog (similar to XIPV)
class XReservesRemainingTablesDialog(QDialog):
    def __init__(self, parent=None, tables=None):
        super().__init__(parent)
        self.setWindowTitle("Remaining XReserves Tables")
        self.setMinimumSize(800, 600)
        self.tables = tables if tables else [None] * 5  # Changed to 5 tables
        self.modified_tables = set()  # Track which tables have been modified
        
        # Main layout
        main_layout = QVBoxLayout()
        
        # Instructions
        instructions = QLabel(
            "Tables 1 and 2 have been imported. Default data is loaded for Tables 3-7.\n"
            "You can view and modify the default data or import new data."
        )
        instructions.setStyleSheet("font-weight: bold;")
        main_layout.addWidget(instructions)
        
        # Create tab widget for tables
        self.tab_widget = QTabWidget()
        
        # Add tabs for tables 3-7 (5 tables)
        for i in range(5):
            tab = self.create_table_tab(i+3, self.tables[i])
            self.tab_widget.addTab(tab, f"Table {i+3}")
        
        main_layout.addWidget(self.tab_widget)
        
        # Buttons
        button_box = QHBoxLayout()
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.ok_button = QPushButton("Finish")
        self.ok_button.clicked.connect(self.accept)
        
        button_box.addWidget(self.cancel_button)
        button_box.addWidget(self.ok_button)
        main_layout.addLayout(button_box)
        
        self.setLayout(main_layout)
    
    def create_table_tab(self, table_num, table_data):
        """Create a tab for a table"""
        tab_widget = QWidget()
        tab_layout = QVBoxLayout()
        
        # Button to import new data
        import_button = QPushButton(f"Import New Data for Table {table_num}")
        import_button.clicked.connect(lambda: self.import_new_data(table_num-3))  # Adjusted index calculation
        tab_layout.addWidget(import_button)
        
        # Table widget to display/edit data
        table_widget = QTableWidget()
        self.populate_table_widget(table_widget, table_data)
        table_widget.itemChanged.connect(lambda: self.handle_table_edit(table_num-3, table_widget))  # Adjusted index calculation
        
        tab_layout.addWidget(QLabel("Default Data (editable):"))
        tab_layout.addWidget(table_widget)
        
        # Store reference to table widget
        setattr(self, f"table{table_num}_widget", table_widget)
        
        tab_widget.setLayout(tab_layout)
        return tab_widget
    
    # def populate_table_widget(self, table_widget, table_data):
    #     """Populate a QTableWidget with data from a polars DataFrame"""
    #     if table_data is None or table_data.is_empty():
    #         table_widget.setRowCount(0)
    #         table_widget.setColumnCount(0)
    #         return
        
    #     # Convert polars to pandas for easier handling
    #     pandas_df = table_data.to_pandas()
        
    #     # Set table dimensions
    #     table_widget.setRowCount(len(pandas_df))
    #     table_widget.setColumnCount(len(pandas_df.columns))
    #     table_widget.setHorizontalHeaderLabels(pandas_df.columns)
        
    #     # Fill data
    #     for i in range(len(pandas_df)):
    #         for j in range(len(pandas_df.columns)):
    #             value = pandas_df.iloc[i, j]
    #             item = QTableWidgetItem(str(value))
    #             table_widget.setItem(i, j, item)
        
    #     # Resize columns to content
    #     table_widget.resizeColumnsToContents()
    
    def populate_table_widget(self, table_widget, table_data):
        """Populate a QTableWidget with data from a polars DataFrame"""
        if table_data is None or table_data.is_empty():
            table_widget.setRowCount(0)
            table_widget.setColumnCount(0)
            return
        
        # Convert polars to pandas for easier handling
        pandas_df = table_data.to_pandas()
        populate_table_with_types(table_widget, pandas_df)


    def import_new_data(self, table_index):
        """Import new data for a table"""
        dialog = XReservesTableDialog(self, table_number=table_index+3)  # Adjusted table number
        if dialog.exec():
            table_data = dialog.table_data
            if table_data:
                try:
                    # Parse the clipboard data into pandas DataFrame
                    pandas_df = pd.read_clipboard(sep='\t')
                    # Convert to polars
                    pl_df = pl.from_pandas(pandas_df)
                    
                    # Update the table
                    self.tables[table_index] = pl_df
                    
                    # Update UI
                    table_widget = getattr(self, f"table{table_index+3}_widget")  # Adjusted attribute name
                    self.populate_table_widget(table_widget, pl_df)
                    
                    # Mark as modified
                    self.modified_tables.add(table_index)
                    
                    QMessageBox.information(self, "Success", f"New data imported for Table {table_index+3}")  # Adjusted table number
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to parse table data: {str(e)}")
    
    # def handle_table_edit(self, table_index, table_widget):
    #     """Handle edits to the table widget"""
    #     try:
    #         # Get headers
    #         headers = []
    #         for j in range(table_widget.columnCount()):
    #             headers.append(table_widget.horizontalHeaderItem(j).text())
            
    #         # Get all data
    #         data = []
    #         for i in range(table_widget.rowCount()):
    #             row_data = []
    #             for j in range(table_widget.columnCount()):
    #                 item = table_widget.item(i, j)
    #                 value = item.text() if item else ""
    #                 row_data.append(value)
    #             data.append(row_data)
            
    #         # Create pandas DataFrame and convert to polars
    #         pandas_df = pd.DataFrame(data, columns=headers)
    #         pl_df = pl.from_pandas(pandas_df)
            
    #         # Update the table
    #         self.tables[table_index] = pl_df
            
    #         # Mark as modified
    #         self.modified_tables.add(table_index)
    #     except Exception as e:
    #         print(f"Error updating table data: {str(e)}")

    def handle_table_edit(self, table_index, table_widget):
        """Handle edits to the table widget preserving types"""
        try:
            # Extract data preserving types
            pandas_df = extract_table_data(table_widget)
            
            # Convert to polars
            pl_df = pl.from_pandas(pandas_df)
            
            # Update the table
            self.tables[table_index] = pl_df
            
            # Mark as modified
            self.modified_tables.add(table_index)
        except Exception as e:
            print(f"Error updating table data: {str(e)}")

# YReserves Window
class YReservesWindow(FramelessWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("YReserves")
        self.setGeometry(200, 200, 600, 400)
        
        # Central widget
        # central_widget = QWidget()
        # self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(self.content_area)
        
        # Title
        title = QLabel("YReserves Processing")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: bold;")
        
        # Process button
        self.process_button = QPushButton("Process")
        self.process_button.clicked.connect(self.process_data)
        
        # Add widgets to layout
        main_layout.addWidget(title)
        main_layout.addWidget(self.process_button)
    
    def process_data(self):
        # ====== ADD YOUR FUNCTION HERE ======
        # Replace this comment with your processing function
        # Example: result = process_yreserves()
        # Currently just showing message as per requirements
        # =====================================
        
        QMessageBox.information(self, "Processing", "YReserves processing initiated")

class FirstFileDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Load First File")
        self.setMinimumWidth(600)
        self.setMinimumHeight(400)
        self.file_path = None
        
        layout = QVBoxLayout()
        
        # File section
        file_group = QGroupBox("Select First File")
        file_layout = QVBoxLayout()
        
        # Drag-drop area
        self.drag_drop = FileDragDropWidget()
        
        # Browse button
        self.browse_button = QPushButton("Browse")
        self.browse_button.clicked.connect(self.browse_file)
        
        file_layout.addWidget(self.drag_drop)
        file_layout.addWidget(self.browse_button)
        file_group.setLayout(file_layout)
        
        # Date input
        date_layout = QHBoxLayout()
        self.date_input = QDateEdit()
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setCalendarPopup(True)
        date_layout.addWidget(QLabel("Date:"))
        date_layout.addWidget(self.date_input)
        
        # Buttons
        button_box = QHBoxLayout()
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.next_button = QPushButton("Next")
        self.next_button.clicked.connect(self.validate_and_accept)
        
        button_box.addWidget(self.cancel_button)
        button_box.addWidget(self.next_button)
        
        # Add all widgets to main layout
        layout.addWidget(file_group)
        layout.addLayout(date_layout)
        layout.addLayout(button_box)
        
        self.setLayout(layout)
    
    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select File", "", "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            self.file_path = file_path
            self.drag_drop.label.setText(f"File: {os.path.basename(file_path)}")
    
    def validate_and_accept(self):
        self.file_path = getattr(self.drag_drop, 'file_path', None)
        
        if not self.file_path:
            QMessageBox.warning(self, "Warning", "Please select a file.")
            return
        
        self.accept()

class SecondFileDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Load Second File")
        self.setMinimumWidth(600)
        self.setMinimumHeight(400)
        self.file_path = None
        
        layout = QVBoxLayout()
        
        # File section
        file_group = QGroupBox("Select Second File")
        file_layout = QVBoxLayout()
        
        # Drag-drop area
        self.drag_drop = FileDragDropWidget()
        
        # Browse button
        self.browse_button = QPushButton("Browse")
        self.browse_button.clicked.connect(self.browse_file)
        
        file_layout.addWidget(self.drag_drop)
        file_layout.addWidget(self.browse_button)
        file_group.setLayout(file_layout)
        
        # Buttons
        button_box = QHBoxLayout()
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.calculate_button = QPushButton("Calculate")
        self.calculate_button.clicked.connect(self.validate_and_accept)
        
        button_box.addWidget(self.cancel_button)
        button_box.addWidget(self.calculate_button)
        
        # Add all widgets to main layout
        layout.addWidget(file_group)
        layout.addLayout(button_box)
        
        self.setLayout(layout)
    
    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select File", "", "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            self.file_path = file_path
            self.drag_drop.label.setText(f"File: {os.path.basename(file_path)}")
    
    def validate_and_accept(self):
        self.file_path = getattr(self.drag_drop, 'file_path', None)
        
        if not self.file_path:
            QMessageBox.warning(self, "Warning", "Please select a file.")
            return
        
        self.accept()

# Other Tool Widgets
class OtherToolsWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        title = QLabel("Other Tools")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        
        self.ccy_segregation_btn = QPushButton("Full Reval CCY Segregation")
        self.currency_pair_btn = QPushButton("Currency Pair Replacement")
        
        self.ccy_segregation_btn.clicked.connect(self.open_ccy_segregation)
        self.currency_pair_btn.clicked.connect(self.open_currency_pair)
        
        layout.addWidget(title)
        layout.addWidget(self.ccy_segregation_btn)
        layout.addWidget(self.currency_pair_btn)
        
        self.setLayout(layout)
    
    def open_ccy_segregation(self):
        self.ccy_window = CCYSegregationWindow()
        self.ccy_window.show()
    
    def open_currency_pair(self):
        self.currency_window = CurrencyPairWindow()
        self.currency_window.show()

class CCYSegregationWindow(FramelessWindow):
    def __init__(self):
        super().__init__()
        self.input_data = {}
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("Full Reval CCY Segregation")
        self.setGeometry(200, 200, 800, 600)
        
        # central_widget = QWidget()
        # self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(self.content_area)
        
        # Process steps
        self.stacked_steps = QStackedWidget()
        
        # Step 1: File inputs
        self.step1 = self.create_file_input_step()
        # Step 2: Table input
        self.step2 = self.create_table_input_step()
        
        self.stacked_steps.addWidget(self.step1)
        self.stacked_steps.addWidget(self.step2)
        
        main_layout.addWidget(self.stacked_steps)
        
    def create_file_input_step(self):
        widget = QWidget()
        layout = QFormLayout()
        
        self.file_inputs = {}
        for i in range(1, 4):
            btn = QPushButton(f"Browse File {i}")
            le = QLineEdit()
            sheet_le = QLineEdit()
            
            btn.clicked.connect(lambda _, n=i: self.browse_file(n))
            self.file_inputs[f"file{i}"] = le
            self.file_inputs[f"sheet{i}"] = sheet_le
            
            file_layout = QHBoxLayout()
            file_layout.addWidget(le)
            file_layout.addWidget(btn)
            
            layout.addRow(f"File {i} Path:", file_layout)
            layout.addRow(f"Sheet {i} Name:", sheet_le)
        
        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        self.calendar.setStyleSheet("""
            QCalendarWidget QToolButton {
                min-width: 80px;
                color: white;
            }
        """)

        self.next_btn = QPushButton("Next")
        self.next_btn.clicked.connect(lambda: self.stacked_steps.setCurrentIndex(1))
        
        layout.addRow("Select Date:", self.calendar)
        layout.addRow(self.next_btn)
        
        widget.setLayout(layout)
        return widget
    
    # def create_table_input_step(self):
    #     widget = QWidget()
    #     layout = QVBoxLayout()
        
    #     self.table_input = QPlainTextEdit()
    #     self.table_input.setPlaceholderText("Paste table data here...")
    #     self.process_btn = QPushButton("Process")
    #     self.process_btn.clicked.connect(self.collect_all_data)
        
    #     layout.addWidget(QLabel("Paste Table Data:"))
    #     layout.addWidget(self.table_input)
    #     layout.addWidget(self.process_btn)
        
    #     widget.setLayout(layout)
    #     return widget

    def create_table_input_step(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Input area
        self.table_input = QPlainTextEdit()
        self.table_input.setPlaceholderText("Paste table data here...")
        self.table_input.textChanged.connect(self.update_preview)
        
        # Preview table
        self.preview_table = QTableWidget()
        self.preview_table.setMaximumHeight(150)
        self.preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        # Process button
        self.process_btn = QPushButton("Process")
        self.process_btn.clicked.connect(self.collect_all_data)
        
        layout.addWidget(QLabel("Paste Table Data:"))
        layout.addWidget(self.table_input)
        layout.addWidget(QLabel("Preview:"))
        layout.addWidget(self.preview_table)
        layout.addWidget(self.process_btn)
        
        widget.setLayout(layout)
        return widget

    def update_preview(self):
        try:
            # Parse clipboard data
            df = pd.read_csv(StringIO(self.table_input.toPlainText()), sep='\t')
            
            # Update preview table
            self.preview_table.setRowCount(min(5, len(df)))
            self.preview_table.setColumnCount(len(df.columns))
            self.preview_table.setHorizontalHeaderLabels(df.columns)
            
            # Fill data
            for i in range(min(5, len(df))):
                for j in range(len(df.columns)):
                    self.preview_table.setItem(i, j, 
                        QTableWidgetItem(str(df.iloc[i, j])))
            
            # Resize columns
            self.preview_table.resizeColumnsToContents()
            
        except Exception as e:
            self.preview_table.clear()
            self.preview_table.setRowCount(0)
            self.preview_table.setColumnCount(0)
    
    
    def browse_file(self, file_num):
        path, _ = QFileDialog.getOpenFileName()
        if path:
            self.file_inputs[f"file{file_num}"].setText(path)
    
    def collect_all_data(self):
        # selected_date = self.calendar.selectedDate().toString("yyyy-MM-dd")
        # Collect file paths and sheets
        self.input_data = {
            "file1": self.file_inputs["file1"].text(),
            "sheet1": self.file_inputs["sheet1"].text(),
            "file2": self.file_inputs["file2"].text(),
            "sheet2": self.file_inputs["sheet2"].text(),
            "file3": self.file_inputs["file3"].text(),
            "sheet3": self.file_inputs["sheet3"].text(),
            #################################################
            "date": self.calendar.selectedDate().toString("yyyy-MM-dd"),
            "table_data": self.parse_table_data()
        }
        
        # Here you would call your processing function
        # process_ccy_segregation(self.input_data)
        
        QMessageBox.information(self, "Success", "All data collected ready for processing")
    
    def parse_table_data(self):
        try:
            return pl.from_pandas(pd.read_clipboard(sep='\t'))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Table parsing failed: {str(e)}")
            return None

class CurrencyPairWindow(FramelessWindow):
    def __init__(self):
        super().__init__()
        self.default_excel = "CCYreplacement.xlsx"
        self.table_data = [None, None]  # Stores both tables' data
        self.init_ui()
        self.load_or_create_default_excel()

    def init_ui(self):
        self.setWindowTitle("Currency Pair Replacement")
        self.setGeometry(200, 200, 800, 600)
        
        # central_widget = QWidget()
        # self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(self.content_area)
        
        # File inputs
        # file_group = QGroupBox("Input Files")
        file_group = QGroupBox()
        file_layout = QFormLayout()
        
        self.file_inputs = []
        for i in range(3):
            le = QLineEdit()
            btn = QPushButton("Browse")
            btn.clicked.connect(lambda _, n=i: self.browse_file(n))
            hbox = QHBoxLayout()
            hbox.addWidget(le)
            hbox.addWidget(btn)
            file_layout.addRow(f"File {i+1}:", hbox)
            self.file_inputs.append(le)
        
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)

        # Table tabs
        self.tabs = QTabWidget()
        self.table_widgets = []
        
        for i in range(2):
            tab = QWidget()
            layout = QVBoxLayout(tab)
            
            # Table widget
            table = QTableWidget()
            table.setEditTriggers(QTableWidget.AllEditTriggers)
            table.cellChanged.connect(lambda row, col, idx=i: self.table_updated(row, col, idx))
            
            # Buttons
            btn_layout = QHBoxLayout()
            load_btn = QPushButton("Reload Default")
            load_btn.clicked.connect(lambda _, idx=i: self.load_table(idx))
            save_btn = QPushButton("Save Changes")
            save_btn.clicked.connect(lambda _, idx=i: self.save_table(idx))
            
            btn_layout.addWidget(load_btn)
            btn_layout.addWidget(save_btn)
            
            layout.addWidget(table)
            layout.addLayout(btn_layout)
            self.table_widgets.append(table)
            self.tabs.addTab(tab, f"Table {i+1}")
        
        main_layout.addWidget(self.tabs)
        
        # Process button
        process_btn = QPushButton("Process Replacement")
        process_btn.clicked.connect(self.process_data)
        main_layout.addWidget(process_btn)

    def load_or_create_default_excel(self):
        """Create or load default Excel file with two sheets"""
        try:
            if not os.path.exists(self.default_excel):
                wb = Workbook()
                wb.remove(wb.active)  # Remove default sheet
                for i in range(2):
                    wb.create_sheet(f"Table{i+1}")
                wb.save(self.default_excel)
            
            self.load_table(0)
            self.load_table(1)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to initialize Excel file: {str(e)}")

    def load_table(self, table_idx):
        """Load table from Excel sheet"""
        try:
            wb = load_workbook(self.default_excel)
            sheet = wb.worksheets[table_idx]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            
            df = pd.DataFrame(data[1:], columns=data[0]) if data else pd.DataFrame()
            self.table_data[table_idx] = df
            self.populate_table(table_idx)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load table {table_idx+1}: {str(e)}")

    # Inside CurrencyPairWindow class
    def populate_table(self, table_idx):
        """Populate QTableWidget with data preserving types"""
        table = self.table_widgets[table_idx]
        df = self.table_data[table_idx]
        
        table.blockSignals(True)
        table.clear()
        populate_table_with_types(table, df)
        table.blockSignals(False)

    # Inside CurrencyPairWindow class
    def table_updated(self, row, col, table_idx):
        """Handle table edits preserving data types"""
        try:
            item = self.table_widgets[table_idx].item(row, col)
            if item is None or not item.text():
                value = ""
            else:
                text = item.text()
                try:
                    # Attempt to convert to number if possible
                    if text.isdigit():
                        value = int(text)
                    elif text.replace(".", "", 1).isdigit():
                        value = float(text)
                    else:
                        value = text
                except ValueError:
                    value = text
            
            col_name = self.table_data[table_idx].columns[col]
            self.table_data[table_idx].at[row, col_name] = value
        except Exception as e:
            QMessageBox.warning(self, "Edit Error", f"Invalid input: {str(e)}")

    def save_table(self, table_idx):
        """Save table back to Excel"""
        try:
            wb = load_workbook(self.default_excel)
            sheet = wb.worksheets[table_idx]
            sheet.delete_rows(1, sheet.max_row)  # Clear existing data
            
            # Write new data
            df = self.table_data[table_idx]
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                sheet.append(row)
            
            wb.save(self.default_excel)
            QMessageBox.information(self, "Success", f"Table {table_idx+1} saved successfully")
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save table: {str(e)}")

    def browse_file(self, index):
        path, _ = QFileDialog.getOpenFileName()
        if path:
            self.file_inputs[index].setText(path)

    def process_data(self):
        """Collect all data for processing"""
        data = {
            'files': [le.text() for le in self.file_inputs],
            'tables': self.table_data
        }
        # Add your processing logic here
        QMessageBox.information(self, "Ready", "All data collected for processing")

        
# Main entry point
if __name__ == "__main__":
    app = QApplication(sys.argv)
    with open("dark_theme.qss", "r") as f:
        app.setStyleSheet(f.read())
    login_window = LoginWindow()
    login_window.show()
    sys.exit(app.exec())